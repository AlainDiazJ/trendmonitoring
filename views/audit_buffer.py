#!/usr/bin/env python3
"""
audit_buffer.py — Auditoría de archivos de prueba (hoja 'Buffer') para Trend Monitoring.

QUÉ HACE
--------
Recorre una carpeta de archivos de prueba (.xlsm / .xls), abre la hoja 'Buffer'
de cada uno, lee la columna A (nombres de parámetro) y verifica cuáles de los
parámetros que SÍ queremos monitorear están presentes y cuáles faltan.

IMPORTANTE — PRIVACIDAD
-----------------------
Este script NO lee, NO guarda y NO muestra los VALORES (columna B). Solo trabaja
con los NOMBRES de los parámetros (columna A) y conteos. El reporte que genera es
seguro de compartir: no contiene datos sensibles del motor.

CÓMO SE USA
-----------
    python audit_buffer.py "C:\\ruta\\a\\carpeta_de_pruebas"

    # opcional: indicar familia (por ahora solo LEAP está cargada)
    python audit_buffer.py "C:\\ruta" --family LEAP

SALIDA
------
- Imprime un resumen en pantalla.
- Escribe 'audit_report.txt' (solo nombres y conteos, sin valores) en la carpeta
  desde donde corres el script.
"""

import sys
import argparse
from pathlib import Path
from collections import Counter, defaultdict

# ---------------------------------------------------------------------------
# Lista de parámetros que QUEREMOS monitorear, por familia de motor.
# Tomada de las macros que ya nos pasaron. NO es la lista de las ~6000 filas:
# es la selección curada que importa para el trend.
# Por ahora solo LEAP está completa (es el piloto). CFM56 se añade después.
# ---------------------------------------------------------------------------
WANTED = {
    "LEAP": [
        "Serial Number", "Point Test Number", "Id point", "Point Type",
        "Point Number", "Description", "TIME", "DATE", "PAMBpsia", "PT10psia",
        "N1", "N2", "N2R25", "TRISECel", "YEOPSLpsid_MIN",
        "YEOPSLCpsig_MC_TO_PRE_MIN_LIM", "T2", "PS10psia", "WF36pph",
        "N2_MI_LIM", "FHV", "YTEOSL", "YEOPSLpsig", "YEOPSLCpsig", "HUM",
        "FNlbf", "YEGTSL", "OQ_A2", "OQ_A80", "YTEOSL_MAX_LIM",
        "YEOPSLCpsig_MC_TO_PRE_MAX_LIM", "N1RKH", "FNRKH", "N1R2rated", "N2R2",
        "EGTR2", "FNR2", "WFMR2", "W2AR2", "N1R2tested", "FNRL", "THD", "N1HD",
        "N2HD", "EGTR2HD", "WFMR2FN", "FMN1", "N1HDM", "N2HDM", "EGTR2HDM",
        "FNR2M", "NAMEPLATETHRUST", "MAX_N2HDM", "MAX_EGTR2HDM", "MIN_FNR2",
        "FNMR", "N2MAR", "EGTMAR", "FNMAR", "FNMARNPT", "PCELLFpsig",
        "PCELLRpsig", "PT10psig", "PT10_1_psig", "PT10_2_psig", "PT10_3_psig",
        "PT10_4_psig", "PS10W1psig", "PS10W2psig", "PS10W3psig", "PS10W4psig",
        "PT154_1_psig", "PT154_2_psig", "PT154_3_psig", "PT154_4_psig", "PS3",
        "TACALR", "TACFUR", "TACAUL", "TACFUL", "TACMUL", "TACAUR", "TACMLR",
        "TACMUR", "TACFLL", "TACALL", "TACMLL", "TACFLR", "T12ENGSEL", "T25SEL",
        "T3SEL", "EGTSEL", "WFMAINpph", "WFVERIFYpph", "DWF", "TWFE", "RH",
        "FSGxxx", "TFSGxx", "P154psia", "FUELFILTERDPSEL", "FUELFILTERDPSEL-A",
        "FUELFILTERDPSEL-B", "FUELFILTERDP", "FUELFILTERDP-A", "FUELFILTERDP-B",
    ],
}

# Nombre esperado de la hoja de datos crudos
BUFFER_SHEET = "Buffer"


def norm(s):
    """Normaliza un nombre para comparar de forma tolerante:
    quita espacios extremos, colapsa espacios internos, ignora mayúsculas."""
    if s is None:
        return ""
    return " ".join(str(s).strip().split()).lower()


def read_buffer_names(path):
    """Devuelve la lista de nombres de la columna A de la hoja Buffer.
    Maneja .xlsm/.xlsx (openpyxl) y .xls legacy (pandas+xlrd).
    NUNCA lee la columna B (valores)."""
    ext = path.suffix.lower()

    if ext in (".xlsm", ".xlsx"):
        from openpyxl import load_workbook
        # read_only=True es clave: las hojas Buffer tienen ~6000 filas
        wb = load_workbook(path, read_only=True, data_only=True)
        # buscar la hoja Buffer de forma tolerante a mayúsculas
        sheet = None
        for name in wb.sheetnames:
            if norm(name) == norm(BUFFER_SHEET):
                sheet = name
                break
        if sheet is None:
            wb.close()
            return None, wb.sheetnames  # no encontró Buffer; devuelve hojas existentes
        ws = wb[sheet]
        names = []
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            names.append(row[0])
        wb.close()
        return names, None

    elif ext == ".xls":
        import pandas as pd
        # Solo columna A (usecols=[0]); nunca tocamos valores
        try:
            xls = pd.ExcelFile(path, engine="xlrd")
        except Exception as e:
            return None, [f"<error abriendo .xls: {e}>"]
        sheet = None
        for name in xls.sheet_names:
            if norm(name) == norm(BUFFER_SHEET):
                sheet = name
                break
        if sheet is None:
            return None, xls.sheet_names
        df = pd.read_excel(xls, sheet_name=sheet, usecols=[0], header=None)
        return df.iloc[:, 0].tolist(), None

    else:
        return None, [f"<extensión no soportada: {ext}>"]


def audit_folder(folder, family):
    folder = Path(folder)
    files = sorted(
        [p for p in folder.iterdir()
         if p.suffix.lower() in (".xlsm", ".xlsx", ".xls") and not p.name.startswith("~")]
    )

    wanted = WANTED.get(family)
    if wanted is None:
        print(f"[!] Familia '{family}' no está definida todavía. Familias disponibles: {list(WANTED)}")
        sys.exit(1)
    wanted_norm = {norm(w): w for w in wanted}

    report = []
    report.append(f"AUDITORÍA DE BUFFER — familia {family}")
    report.append(f"Carpeta: {folder}")
    report.append(f"Archivos encontrados: {len(files)}")
    report.append(f"Parámetros buscados (de la macro): {len(wanted)}")
    report.append("=" * 70)

    # Conteo global: en cuántos archivos aparece cada parámetro buscado
    found_in_n = Counter()
    # Nombres en el Buffer que NO están en nuestra lista (candidatos a revisar)
    extras_seen = Counter()
    per_file = {}

    for f in files:
        names, problem = read_buffer_names(f)
        if names is None:
            per_file[f.name] = {"ok": False, "msg": f"No se halló hoja 'Buffer'. Hojas: {problem}"}
            continue

        present_norm = set()
        total_rows = 0
        for nm in names:
            if nm is None or str(nm).strip() == "":
                continue
            total_rows += 1
            nn = norm(nm)
            present_norm.add(nn)

        found_here = [w for wn, w in wanted_norm.items() if wn in present_norm]
        missing_here = [w for wn, w in wanted_norm.items() if wn not in present_norm]
        for w in found_here:
            found_in_n[w] += 1

        # nombres del buffer que no buscamos (solo cuenta, para no inundar)
        for nn in present_norm:
            if nn not in wanted_norm:
                extras_seen[nn] += 1

        per_file[f.name] = {
            "ok": True,
            "total_rows": total_rows,
            "found": len(found_here),
            "missing": missing_here,
        }

    # ---- Reporte por archivo ----
    report.append("\nPOR ARCHIVO:")
    for fname, info in per_file.items():
        if not info["ok"]:
            report.append(f"  [ERROR] {fname}: {info['msg']}")
            continue
        report.append(
            f"  [OK] {fname}: {info['total_rows']} filas en Buffer, "
            f"{info['found']}/{len(wanted)} parámetros buscados encontrados"
        )
        if info["missing"]:
            report.append(f"        Faltan: {', '.join(info['missing'])}")

    # ---- Resumen de cobertura ----
    ok_files = [f for f, i in per_file.items() if i["ok"]]
    report.append("\nRESUMEN DE COBERTURA (en cuántos archivos aparece cada parámetro buscado):")
    if ok_files:
        always = [w for w in wanted if found_in_n[w] == len(ok_files)]
        sometimes = [w for w in wanted if 0 < found_in_n[w] < len(ok_files)]
        never = [w for w in wanted if found_in_n[w] == 0]
        report.append(f"  Presentes en TODOS los archivos ({len(always)}): {', '.join(always) if always else '—'}")
        report.append(f"  Presentes en ALGUNOS ({len(sometimes)}): {', '.join(sometimes) if sometimes else '—'}")
        report.append(f"  Presentes en NINGUNO ({len(never)}): {', '.join(never) if never else '—'}")
        report.append("\n  -> Revisa los de 'NINGUNO': o el nombre está escrito distinto en el Buffer,")
        report.append("     o ese parámetro no existe en estos archivos.")

    # ---- Extras (nombres en Buffer que no buscamos) ----
    report.append(f"\nNOMBRES EN EL BUFFER QUE NO ESTAMOS BUSCANDO: {len(extras_seen)} distintos.")
    report.append("  (No se listan todos para no inundar el reporte; son en su mayoría parámetros de setup.)")
    report.append("  Si crees que falta algún parámetro útil para el trend, dime y lo añadimos a la lista.")

    text = "\n".join(report)
    print(text)

    out = Path("audit_report.txt")
    out.write_text(text, encoding="utf-8")
    print(f"\n[Reporte guardado en: {out.resolve()}]")
    print("[Recuerda: este reporte solo contiene NOMBRES y conteos, ningún valor sensible.]")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Auditoría de la hoja Buffer para Trend Monitoring.")
    ap.add_argument("folder", help="Carpeta con archivos de prueba (.xlsm/.xls)")
    ap.add_argument("--family", default="LEAP", help="Familia de motor (por ahora: LEAP)")
    args = ap.parse_args()
    audit_folder(args.folder, args.family)
