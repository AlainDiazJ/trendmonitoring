#!/usr/bin/env python3
"""
extract_units.py — Extrae NOMBRE (col A) + UNIDAD (col C) de la hoja 'Buffer'.

QUÉ HACE
--------
Para los parámetros que nos interesan (lista por variante), lee de cada archivo
el nombre del parámetro (columna A) y su unidad (columna C), en la misma fila.
Genera un catálogo nombre->unidad, y avisa si la unidad cambia entre archivos.

IMPORTANTE — PRIVACIDAD
-----------------------
Lee SOLO columnas A y C (nombre y unidad). NUNCA lee la columna B (valores
sensibles del motor). El reporte generado es seguro de compartir.

CÓMO SE USA
-----------
    py extract_units.py "C:\\pruebas_leap_1B" --variant 1B
    py extract_units.py "C:\\pruebas_leap_1A" --variant 1A

SALIDA
------
- Imprime el catálogo en pantalla.
- Escribe 'units_<variant>.txt' (nombre + unidad + cobertura), sin valores.
"""

import sys
import argparse
from pathlib import Path
from collections import defaultdict

BUFFER_SHEET = "Buffer"

# Parámetros de interés por variante. Incluye el núcleo + los EGT solicitados.
# Si un nombre no se encuentra, el reporte lo marcará (igual que la auditoría).
WANTED = {
    "1B": [
        # identificación
        "Serial Number", "Point Test Number", "Id point", "Point Type",
        "Point Number", "Description", "TIME", "DATE",
        # ambientales (ISA)
        "PAMBpsia", "T2", "HUM", "RH",
        # motor primario
        "N1", "N2", "N2R25", "WF36pph", "FNlbf",
        # EGT (cadena R2 + crudo/margen solicitados)
        "EGT", "EGTK", "EGTM", "EGTR2", "EGTR2HD", "EGTR2HDM",
        # salud de celda
        "PCELLFpsig", "PCELLRpsig",
        # combustible
        "FHV",
    ],
    "1A": [
        # identificación
        "Serial Number", "Point Test Number", "Id point", "Point Type",
        "POINT_TYPE_N", "Point Number", "Description", "DATE",
        # ambientales (ISA) — nombres tipicos de 1A
        "PAMB", "PAMBkPa", "T2", "HUM", "HUM4",
        # motor primario — cadena K
        "N1", "N2", "N1K", "N2K", "WF36", "WF36kgh", "FN", "FNK", "FNdaN",
        # EGT (cadena K + crudo solicitado)
        "EGT", "EGTMEAS", "EGTK", "EGTK1", "EGTK2", "EGTK3", "EGTK3M",
        "EGTHD", "EGTHDM", "EGTHD_MAR", "EGTHDM_MAR",
        # salud de celda — nombres tipicos de 1A
        "FPCELL", "RPCELL", "FPCELLkPaa", "RPCELLkPaa",
        # combustible
        "FHV",
    ],
}


def norm(s):
    if s is None:
        return ""
    return " ".join(str(s).strip().split()).lower()


def read_name_unit_rows(path):
    """Devuelve lista de (nombre, unidad) leyendo columnas A y C. Ignora B."""
    ext = path.suffix.lower()
    if ext in (".xlsm", ".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = next((n for n in wb.sheetnames if norm(n) == norm(BUFFER_SHEET)), None)
        if sheet is None:
            wb.close()
            return None
        ws = wb[sheet]
        rows = []
        # min_col=1, max_col=3 -> trae A, B, C. Usamos solo [0] (A) y [2] (C),
        # nunca [1] (B). openpyxl no nos deja saltar la B sin leerla, pero
        # NO la guardamos ni la procesamos en ningún momento.
        for r in ws.iter_rows(min_col=1, max_col=3, values_only=True):
            name = r[0]
            unit = r[2] if len(r) >= 3 else None
            rows.append((name, unit))
        wb.close()
        return rows
    elif ext == ".xls":
        import pandas as pd
        xls = pd.ExcelFile(path, engine="xlrd")
        sheet = next((n for n in xls.sheet_names if norm(n) == norm(BUFFER_SHEET)), None)
        if sheet is None:
            return None
        # usecols=[0,2] -> solo columnas A y C, jamás la B
        df = pd.read_excel(xls, sheet_name=sheet, usecols=[0, 2], header=None,
                           names=["name", "unit"])
        return list(zip(df["name"].tolist(), df["unit"].tolist()))
    return None


def extract(folder, variant):
    folder = Path(folder)
    files = sorted(
        p for p in folder.iterdir()
        if p.suffix.lower() in (".xlsm", ".xlsx", ".xls") and not p.name.startswith("~")
    )
    if not files:
        print(f"[!] No hay archivos en {folder}")
        sys.exit(1)

    wanted = WANTED.get(variant)
    if wanted is None:
        print(f"[!] Variante '{variant}' no definida. Opciones: {list(WANTED)}")
        sys.exit(1)
    wanted_norm = {norm(w): w for w in wanted}

    # para cada parámetro buscado: conjunto de unidades vistas y en cuántos archivos
    units_seen = defaultdict(set)        # nombre canónico -> {unidades}
    found_in_n = defaultdict(int)        # nombre canónico -> nº archivos donde aparece
    n_ok = 0

    for f in files:
        rows = read_name_unit_rows(f)
        if rows is None:
            print(f"  [ERROR] {f.name}: no se encontró hoja 'Buffer'")
            continue
        n_ok += 1
        # mapear nombre normalizado -> unidad en este archivo
        local = {}
        for name, unit in rows:
            if name is None or str(name).strip() == "":
                continue
            nn = norm(name)
            if nn in wanted_norm and nn not in local:
                u = "" if unit is None else str(unit).strip()
                local[nn] = u
        for nn, u in local.items():
            canon = wanted_norm[nn]
            found_in_n[canon] += 1
            units_seen[canon].add(u if u != "" else "(vacío)")

    # ---- Reporte ----
    out = []
    out.append(f"CATÁLOGO NOMBRE+UNIDAD — variante LEAP-{variant}")
    out.append(f"Carpeta: {folder}")
    out.append(f"Archivos leídos OK: {n_ok}/{len(files)}")
    out.append(f"Parámetros buscados: {len(wanted)}")
    out.append("=" * 70)

    out.append("\nENCONTRADOS (nombre -> unidad | en cuántos archivos):")
    for w in wanted:
        if found_in_n[w] > 0:
            us = " / ".join(sorted(units_seen[w]))
            flag = ""
            if len(units_seen[w]) > 1:
                flag = "   <-- ¡UNIDAD INCONSISTENTE entre archivos!"
            out.append(f"  {w:18s} -> {us:20s} | {found_in_n[w]}/{n_ok}{flag}")

    out.append("\nNO ENCONTRADOS (nombre distinto en este formato, o no existe):")
    missing = [w for w in wanted if found_in_n[w] == 0]
    if missing:
        for w in missing:
            out.append(f"  {w}")
    else:
        out.append("  (ninguno — todos encontrados)")

    text = "\n".join(out)
    print(text)
    outfile = Path(f"units_{variant}.txt")
    outfile.write_text(text, encoding="utf-8")
    print(f"\n[Reporte guardado en: {outfile.resolve()}]")
    print("[Solo contiene nombres y unidades. Ningún valor sensible.]")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("folder", help="Carpeta con archivos de una variante")
    ap.add_argument("--variant", required=True, choices=["1A", "1B"],
                    help="Variante LEAP: 1A o 1B")
    args = ap.parse_args()
    extract(args.folder, args.variant)
