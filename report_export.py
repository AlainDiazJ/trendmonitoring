#!/usr/bin/env python3
"""
report_export.py - Genera reportes (Excel y PDF) para Trend Monitoring.

Cada reporte incluye:
  - Datos filtrados (la tabla que se ve en el dashboard)
  - Configuracion/filtros activos del dashboard
  - Estadisticas (media, sigma, limites, outliers) si se proveen
  - La grafica de tendencia como imagen (generada con matplotlib, sin navegador)

No depende de Chrome/kaleido: la imagen se hace con matplotlib.

Lo usa app.py para los botones de descarga.
"""

import html
import io

import matplotlib
matplotlib.use("Agg")  # backend sin ventana, no requiere navegador
import matplotlib.pyplot as plt


def _fmt_si_no(valor):
    return "Si" if bool(valor) else "No"


def grafica_tendencia_png(sub, p_sel, sel_lbl, stats=None, eventos_pos=None,
                          regresion=None, umbrales=None, media_movil=None):
    """Genera la grafica de tendencia como PNG (bytes).

    sub: DataFrame con columnas 'consecutivo' y 'value' (ya filtrado, un parametro).
    stats: dict opcional con media, ucl, lcl, n_sigma.
    eventos_pos: lista opcional de (posicion_x, nombre[, fecha]) para lineas verticales.
    regresion: dict opcional con x/y o pendiente/interseccion/x_min/x_max.
    umbrales: dict opcional con low/high.
    media_movil: dict opcional con x/y/ventana.
    """
    fig, ax = plt.subplots(figsize=(10, 5.2))
    s = sub.sort_values("consecutivo")
    x = s["consecutivo"].astype(float)
    y = s["value"].astype(float)
    ax.plot(x, y, marker="o", linewidth=1.6, color="#1f4e79", label=p_sel)

    if stats:
        media = stats.get("media")
        ucl = stats.get("ucl")
        lcl = stats.get("lcl")
        ns = stats.get("n_sigma")
        if media is not None:
            ax.axhline(media, color="green", linewidth=1.2, label=f"media={media:.2f}")
        if ucl is not None:
            ax.axhline(ucl, color="red", linestyle="--", linewidth=1.1, label=f"+{ns} sigma")
        if lcl is not None:
            ax.axhline(lcl, color="red", linestyle="--", linewidth=1.1, label=f"-{ns} sigma")
        if ucl is not None and lcl is not None:
            ax.fill_between(x, lcl, ucl, color="red", alpha=0.05)

    if umbrales:
        high = umbrales.get("high")
        low = umbrales.get("low")
        if high is not None:
            ax.axhline(high, color="purple", linestyle="-.", linewidth=1.1,
                       label=f"limite sup={high:.2f}")
        if low is not None:
            ax.axhline(low, color="purple", linestyle="-.", linewidth=1.1,
                       label=f"limite inf={low:.2f}")

    if media_movil:
        mx = media_movil.get("x")
        my = media_movil.get("y")
        ventana = media_movil.get("ventana")
        if mx is not None and my is not None:
            ax.plot(mx, my, color="darkorange", linewidth=2.0,
                    label=f"Media movil ({ventana})")

    if regresion:
        rx = regresion.get("x")
        ry = regresion.get("y")
        if rx is None or ry is None:
            pendiente = regresion.get("pendiente")
            interseccion = regresion.get("interseccion")
            x_min = regresion.get("x_min")
            x_max = regresion.get("x_max")
            if None not in (pendiente, interseccion, x_min, x_max):
                rx = [x_min, x_max]
                ry = [pendiente * x_min + interseccion,
                      pendiente * x_max + interseccion]
        if rx is not None and ry is not None:
            label = "Regresion"
            if regresion.get("pendiente") is not None:
                label += f" ({regresion['pendiente']:+.3f}/reporte)"
            ax.plot(rx, ry, color="orange", linestyle=":", linewidth=2.0,
                    label=label)

    if eventos_pos:
        for ev in eventos_pos:
            pos_x = ev[0]
            nombre = ev[1]
            ax.axvline(pos_x, color="teal", linestyle=":", linewidth=1.5)
            ax.text(pos_x, ax.get_ylim()[1], str(nombre), rotation=90,
                    va="top", ha="left", color="teal", fontsize=8)

    ax.set_xlabel("No de reporte (consecutivo)")
    ax.set_ylabel(p_sel)
    ax.set_title(f"Tendencia {p_sel} - {sel_lbl}")
    ax.legend(loc="best", fontsize=8)
    ax.grid(True, alpha=0.3)
    fig.autofmt_xdate(rotation=0)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def grafica_comparacion_png(base, sel_multi, sel_lbl, colores):
    """Genera la grafica de modo comparacion como PNG (bytes), un parametro
    por serie, todas normalizadas 0-100% en un solo eje para que se puedan
    leer juntas sin importar la magnitud. Los valores reales van en la tabla
    del Excel, no en esta imagen.

    base: DataFrame largo con columnas 'consecutivo', 'param_label', 'value'.
    sel_multi: lista de param_label en el orden a graficar/leyenda.
    colores: lista de colores (hex) en el mismo orden que sel_multi, para
    que la leyenda coincida con la grafica en pantalla (Plotly).
    """
    fig, ax = plt.subplots(figsize=(10, 5.2))
    for i, p in enumerate(sel_multi):
        s = base[base["param_label"] == p].sort_values("consecutivo")
        if s.empty:
            continue
        x = s["consecutivo"].astype(float)
        v = s["value"].astype(float)
        rng = v.max() - v.min()
        vn = (v - v.min()) / rng * 100 if rng != 0 else v * 0 + 50
        ax.plot(x, vn, marker="o", linewidth=1.6,
                color=colores[i % len(colores)], label=p)

    ax.set_xlabel("No de reporte (consecutivo)")
    ax.set_ylabel("Valor normalizado (0-100%)")
    ax.set_title(f"Comparacion {sel_lbl} (normalizado 0-100%, ver valores reales en la tabla)")
    ax.legend(loc="best", fontsize=8)
    ax.grid(True, alpha=0.3)
    fig.autofmt_xdate(rotation=0)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def exportar_excel(datos_df, stats_dict, meta, config=None, grafica_png=None):
    """Genera un .xlsx con una hoja Reporte equivalente al PDF.

    Incluye portada con informacion general, filtros/controles, grafica,
    estadisticas y una vista resumida de datos. Las hojas Datos, Filtros e
    Info se mantienen para analisis y traen la tabla completa.
    """
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    azul = "1F4E79"
    azul_claro = "F0F4F8"
    blanco = "FFFFFF"
    borde = Side(style="thin", color="A6A6A6")
    header_fill = PatternFill("solid", fgColor=azul)
    band_fill = PatternFill("solid", fgColor=azul_claro)
    title_font = Font(bold=True, size=16, color=azul)
    section_font = Font(bold=True, size=12, color=azul)
    header_font = Font(bold=True, color=blanco)

    def _write_kv(ws, row, titulo, data, col1=1, col2=2):
        ws.cell(row=row, column=col1, value=titulo).font = section_font
        row += 1
        for k, v in data.items():
            c1 = ws.cell(row=row, column=col1, value=str(k))
            c2 = ws.cell(row=row, column=col2, value=str(v))
            c1.font = Font(bold=True)
            c1.fill = band_fill
            c1.border = c2.border = Border(left=borde, right=borde, top=borde, bottom=borde)
            c2.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        return row + 1

    def _write_table(ws, row, titulo, rows, widths=None):
        ws.cell(row=row, column=1, value=titulo).font = section_font
        row += 1
        if not rows:
            rows = [["Sin datos", "No aplica"]]
        for r_idx, values in enumerate(rows):
            for c_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row + r_idx, column=c_idx, value=value)
                cell.border = Border(left=borde, right=borde, top=borde, bottom=borde)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if r_idx == 0:
                    cell.fill = header_fill
                    cell.font = header_font
                elif r_idx % 2 == 1:
                    cell.fill = band_fill
        if widths:
            for i, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
        return row + len(rows) + 2

    wb = Workbook()

    # Hoja Reporte: equivalente visual del PDF.
    ws_rep = wb.active
    ws_rep.title = "Reporte"
    ws_rep.sheet_view.showGridLines = False
    ws_rep["A1"] = "Reporte de Trend Monitoring"
    ws_rep["A1"].font = title_font
    ws_rep.merge_cells("A1:H1")
    ws_rep.column_dimensions["A"].width = 28
    ws_rep.column_dimensions["B"].width = 42
    for col in range(3, 9):
        ws_rep.column_dimensions[get_column_letter(col)].width = 14

    fila = 3
    fila = _write_kv(ws_rep, fila, "Informacion general", meta)
    fila = _write_kv(ws_rep, fila, "Filtros y controles usados", config or {})

    if grafica_png:
        ws_rep.cell(row=fila, column=1, value="Grafica").font = section_font
        fila += 1
        img = XLImage(io.BytesIO(grafica_png))
        # Mantiene una proporcion parecida al PDF sin hacer enorme el workbook.
        img.width = 850
        img.height = 450
        ws_rep.add_image(img, f"A{fila}")
        fila += 24

    stat_rows = [["Metrica", "Valor"]] + [[str(k), str(v)] for k, v in stats_dict.items()]
    if len(stat_rows) == 1:
        stat_rows.append(["Sin estadisticas", "No aplican con los filtros actuales"])
    fila = _write_table(ws_rep, fila, "Estadisticas", stat_rows, widths=[34, 28])

    max_filas = 60
    cols = list(datos_df.columns)
    data_rows = [cols] + datos_df.head(max_filas).astype(str).values.tolist()
    fila = _write_table(ws_rep, fila, "Datos filtrados", data_rows)
    if len(datos_df) > max_filas:
        ws_rep.cell(
            row=fila,
            column=1,
            value=f"Mostrando {max_filas} de {len(datos_df)} filas. La hoja Datos trae todas.",
        ).font = Font(italic=True, color="666666")

    # Hoja Info
    ws_info = wb.create_sheet("Info")
    ws_info["A1"] = "Reporte Trend Monitoring"
    ws_info["A1"].font = title_font
    fila = 3
    for k, v in meta.items():
        ws_info.cell(row=fila, column=1, value=k).font = Font(bold=True)
        ws_info.cell(row=fila, column=2, value=str(v))
        fila += 1
    ws_info.column_dimensions["A"].width = 28
    ws_info.column_dimensions["B"].width = 70

    # Hoja Filtros / configuracion
    ws_cfg = wb.create_sheet("Filtros")
    ws_cfg.append(["Filtro / control", "Valor"])
    for cell in ws_cfg[1]:
        cell.fill = header_fill
        cell.font = header_font
    for k, v in (config or {}).items():
        ws_cfg.append([k, str(v)])
    ws_cfg.column_dimensions["A"].width = 34
    ws_cfg.column_dimensions["B"].width = 90
    for row in ws_cfg.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Hoja Datos completa
    ws_d = wb.create_sheet("Datos")
    for r in dataframe_to_rows(datos_df, index=False, header=True):
        ws_d.append(r)
    for cell in ws_d[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws_d.freeze_panes = "A2"
    ws_d.auto_filter.ref = ws_d.dimensions
    for col in range(1, len(datos_df.columns) + 1):
        ws_d.column_dimensions[get_column_letter(col)].width = 18

    # Hoja Estadisticas
    ws_s = wb.create_sheet("Estadisticas")
    ws_s.append(["Metrica", "Valor"])
    for cell in ws_s[1]:
        cell.fill = header_fill
        cell.font = header_font
    for k, v in stats_dict.items():
        ws_s.append([k, v])
    if not stats_dict:
        ws_s.append(["Sin estadisticas", "No aplican con los filtros actuales"])
    ws_s.column_dimensions["A"].width = 34
    ws_s.column_dimensions["B"].width = 28

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

def exportar_pdf(datos_df, stats_dict, meta, grafica_png, config=None):
    """Genera un .pdf con encabezado, filtros, grafica, estadisticas y tabla."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image,
                                    Table, TableStyle)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph("Reporte de Trend Monitoring", styles["Title"]))
    elems.append(Spacer(1, 0.3*cm))

    # Info / meta
    elems.append(Paragraph("Informacion general", styles["Heading2"]))
    for k, v in meta.items():
        elems.append(Paragraph(f"<b>{html.escape(str(k))}:</b> {html.escape(str(v))}", styles["Normal"]))
    elems.append(Spacer(1, 0.35*cm))

    # Filtros / configuracion
    if config:
        elems.append(Paragraph("Filtros y controles usados", styles["Heading2"]))
        cfg_rows = [["Filtro / control", "Valor"]] + [
            [html.escape(str(k)), html.escape(str(v))] for k, v in config.items()
        ]
        t_cfg = Table(cfg_rows, colWidths=[5.2*cm, 11.4*cm], hAlign="LEFT", repeatRows=1)
        t_cfg.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
        ]))
        elems.append(t_cfg)
        elems.append(Spacer(1, 0.35*cm))

    # Grafica
    if grafica_png:
        img = Image(io.BytesIO(grafica_png), width=17*cm, height=9*cm)
        elems.append(img)
        elems.append(Spacer(1, 0.4*cm))

    # Estadisticas
    elems.append(Paragraph("Estadisticas", styles["Heading2"]))
    stat_rows = [["Metrica", "Valor"]] + [[str(k), str(v)] for k, v in stats_dict.items()]
    if len(stat_rows) == 1:
        stat_rows.append(["Sin estadisticas", "No aplican con los filtros actuales"])
    t_stats = Table(stat_rows, hAlign="LEFT")
    t_stats.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elems.append(t_stats)
    elems.append(Spacer(1, 0.4*cm))

    # Tabla de datos (limitada para no hacer un PDF gigante)
    elems.append(Paragraph("Datos filtrados", styles["Heading2"]))
    max_filas = 60
    cols = list(datos_df.columns)
    data_rows = [cols] + datos_df.head(max_filas).astype(str).values.tolist()
    t_data = Table(data_rows, hAlign="LEFT", repeatRows=1)
    t_data.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
    ]))
    elems.append(t_data)
    if len(datos_df) > max_filas:
        elems.append(Spacer(1, 0.2*cm))
        elems.append(Paragraph(
            f"(Mostrando {max_filas} de {len(datos_df)} filas. El Excel trae todas.)",
            ParagraphStyle("nota", parent=styles["Normal"], fontSize=8,
                           textColor=colors.grey)))

    doc.build(elems)
    out.seek(0)
    return out.getvalue()