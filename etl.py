#!/usr/bin/env python3
"""
etl.py â€” ETL del piloto Trend Monitoring (LEAP 1A / 1B).

FLUJO
-----
1. Lee mapping.yaml (diccionario de nombres crudos -> canÃ³nicos).
2. Recorre una carpeta de archivos .xlsm de prueba (cada archivo = un punto).
3. Para cada archivo (.xls, .xlsx o .xlsm):
   - detecta la variante (1A / 1B),
   - lee la hoja 'Buffer' (columnas A nombre, B valor, C unidad),
   - extrae los campos de identidad y las mediciones del nÃºcleo,
   - inserta en SQLite: engines / tests / test_points / measurements.
4. Idempotente: si un punto ya fue cargado (mismo hash), no lo duplica.
5. Registra en una tabla 'ingest_log' quÃ© archivo se cargÃ³ y con quÃ© resultado.

PRIVACIDAD
----------
Los VALORES (columna B) SÃ se cargan a la base local (es su propÃ³sito), pero
la base vive solo en esta mÃ¡quina. Nada se sube a ningÃºn lado.

USO
---
    py etl.py "C:\\pruebas_leap_1B" --db data\\motores.db --mapping mapping.yaml
    py etl.py "C:\\pruebas_leap_1A" --db data\\motores.db --mapping mapping.yaml

Puedes correrlo varias veces y sobre varias carpetas; acumula sin duplicar.
"""

import sys
import argparse
import hashlib
from pathlib import Path
from datetime import date, datetime, time, timezone

import yaml
from sqlalchemy import (
    create_engine, Column, Integer, String, Float, Text, DateTime,
    ForeignKey, UniqueConstraint, select,
)
from sqlalchemy.orm import declarative_base, relationship, Session

Base = declarative_base()
BUFFER_SHEET = "Buffer"


# ============================================================================
# ESQUEMA (modelo de 3 niveles: engine -> test -> point -> measurements)
# ============================================================================
class Engine(Base):
    __tablename__ = "engines"
    id = Column(Integer, primary_key=True)
    serial_number = Column(String, unique=True, nullable=False)
    engine_type = Column(String)            # LEAP-1A / LEAP-1B / ...
    points = relationship("TestPoint", back_populates="engine")


class TestPoint(Base):
    __tablename__ = "test_points"
    id = Column(Integer, primary_key=True)
    engine_id = Column(Integer, ForeignKey("engines.id"), nullable=False)
    variant = Column(String)                # 1A / 1B
    serial_number = Column(String)
    point_test_number = Column(String)
    id_point = Column(String)
    point_type = Column(String)             # TKO / MC / ... (clave para el trend)
    point_number = Column(String)
    description = Column(Text)
    test_date = Column(String)              # compatibilidad: fecha cruda como texto
    test_time = Column(String)              # compatibilidad: hora cruda como texto
    test_date_raw = Column(String)          # valor original del Excel
    test_date_iso = Column(String)          # YYYY-MM-DD cuando se pudo parsear
    test_datetime_iso = Column(String)      # YYYY-MM-DD HH:MM:SS cuando se pudo parsear
    date_parse_status = Column(String)      # ok / ambiguous / missing / error
    date_parse_rule = Column(String)        # regla aplicada para trazabilidad
    source_file = Column(String)            # trazabilidad
    row_hash = Column(String, unique=True)  # idempotencia
    stable_point_key = Column(String, unique=True)  # llave estable para config externa
    ingested_at = Column(DateTime)
    engine = relationship("Engine", back_populates="points")
    measurements = relationship("Measurement", back_populates="point")


class Measurement(Base):
    __tablename__ = "measurements"
    id = Column(Integer, primary_key=True)
    point_id = Column(Integer, ForeignKey("test_points.id"), nullable=False)
    canonical = Column(String)              # p_amb / n1 / egt / ...
    raw_name = Column(String)               # PAMBpsia / PAMB / ...
    value = Column(Float)
    unit = Column(String)
    point = relationship("TestPoint", back_populates="measurements")


class IngestLog(Base):
    __tablename__ = "ingest_log"
    id = Column(Integer, primary_key=True)
    source_file = Column(String)
    status = Column(String)                 # ok / skipped / error
    detail = Column(Text)
    ingested_at = Column(DateTime)


# ============================================================================
# UTILIDADES
# ============================================================================
def norm(s):
    if s is None:
        return ""
    return " ".join(str(s).strip().split()).lower()


def detect_variant(filename, buf=None):
    """Detecta variante/familia soportada a partir del nombre o del Buffer.

    Primero usa senales explicitas del nombre/campos principales. Despues usa
    flags del Buffer. Esto evita que un archivo 7B caiga en 5A solo porque el
    Buffer contiene tambien textos genericos como FCS_CFM56-5 configuration.
    """
    def _clean(x):
        return str(x).upper().replace("_", "-").replace(" ", "")

    principales = [str(filename)]
    flags = []
    if buf:
        for raw in (
            "Model Name", "EngineName", "Engine", "EngineType",
            "EngineVersion", "CURRENT_ENGINE_TYPE", "ENGINETYPE",
        ):
            val, _unit = buf.get(norm(raw), (None, None))
            if val is not None:
                principales.append(str(val))
        for key, (val, _unit) in buf.items():
            if "cfm56" in key:
                flags.append(key)
                if val not in (None, ""):
                    flags.append(str(val))

    txt_principal = _clean(" ".join(principales))
    txt_flags = _clean(" ".join(flags))

    # Si el archivo/campos principales dicen LEAP explicitamente, eso manda.
    # Algunos buffers LEAP contienen textos auxiliares CFM, y antes eso podia
    # hacer que un LEAP1A restaurado cayera como CFM56-7B.
    if "LEAP1A" in txt_principal or "LEAP-1A" in txt_principal:
        return "1A"
    if "LEAP1B" in txt_principal or "LEAP-1B" in txt_principal:
        return "1B"

    # Prioridad 7B antes que 5A: algunos buffers 7B contienen ambos flags.
    if "CFM56-7B" in txt_principal or "CFM567B" in txt_principal or "-7B" in txt_principal or "7B" in txt_principal:
        return "CFM56-7B"
    if "CFM56-5A" in txt_principal or "CFM565A" in txt_principal or "-5A" in txt_principal or "5A" in txt_principal:
        return "CFM56-5A"
    if "CFM56-7" in txt_principal or "CFM567" in txt_principal:
        return "CFM56-7B"
    if "CFM56-5" in txt_principal or "CFM565" in txt_principal:
        return "CFM56-5A"

    if "CFM56-7B" in txt_flags or "CFM567B" in txt_flags or "CFM56-7" in txt_flags or "CFM567" in txt_flags:
        return "CFM56-7B"
    if "CFM56-5A" in txt_flags or "CFM565A" in txt_flags or "CFM56-5" in txt_flags or "CFM565" in txt_flags:
        return "CFM56-5A"

    if "LEAP1A" in txt_principal or "LEAP-1A" in txt_principal or ("1A" in txt_principal and "1B" not in txt_principal):
        return "1A"
    if "LEAP1B" in txt_principal or "LEAP-1B" in txt_principal or "1B" in txt_principal:
        return "1B"
    return None

def display_engine_type(variant):
    if variant in ("1A", "1B"):
        return f"LEAP-{variant}"
    return variant


def read_buffer(path):
    """Lee la hoja Buffer y devuelve dict {nombre_normalizado: (valor, unidad)}.
    Lee columnas A (nombre), B (valor), C (unidad).

    Soporta .xlsx/.xlsm con openpyxl y .xls (Excel 97-2003) con pandas/xlrd.
    """
    path = Path(path)
    out = {}

    if path.suffix.lower() == ".xls":
        import pandas as pd
        xls = pd.ExcelFile(path, engine="xlrd")
        sheet = next((n for n in xls.sheet_names if norm(n) == norm(BUFFER_SHEET)), None)
        if sheet is None:
            return None
        raw = pd.read_excel(xls, sheet_name=sheet, header=None, usecols=[0, 1, 2])
        for row in raw.itertuples(index=False, name=None):
            name = row[0] if len(row) >= 1 else None
            if name is None or str(name).strip() == "" or pd.isna(name):
                continue
            nn = norm(name)
            if nn in out:
                continue  # primera aparicion gana
            value = row[1] if len(row) >= 2 and not pd.isna(row[1]) else None
            unit = row[2] if len(row) >= 3 and not pd.isna(row[2]) else None
            out[nn] = (value, unit)
        return out

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = next((n for n in wb.sheetnames if norm(n) == norm(BUFFER_SHEET)), None)
    if sheet is None:
        wb.close()
        return None
    ws = wb[sheet]
    for r in ws.iter_rows(min_col=1, max_col=3, values_only=True):
        name = r[0]
        if name is None or str(name).strip() == "":
            continue
        nn = norm(name)
        if nn in out:
            continue  # primera aparicion gana
        value = r[1] if len(r) >= 2 else None
        unit = r[2] if len(r) >= 3 else None
        out[nn] = (value, unit)
    wb.close()
    return out


def to_float(v):
    """Convierte a float si se puede; si no, None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return None


def _clean_raw(v):
    if v is None:
        return None
    if isinstance(v, float) and pd_isna(v):
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.isoformat()
    txt = str(v).strip()
    return txt or None


def pd_isna(v):
    try:
        import pandas as pd
        return bool(pd.isna(v))
    except Exception:
        return False


def parse_excel_date_time(date_value, time_value=None):
    """Normaliza fecha/hora del Buffer a ISO.

    Preferimos DD/MM/YYYY para textos con slash porque ese es el formato usado
    historicamente por la app. Si dia y mes son <= 12 se marca ambiguous, pero
    se conserva la interpretacion DD/MM para no romper el orden existente.
    """
    raw_date = _clean_raw(date_value)
    raw_time = _clean_raw(time_value)
    if not raw_date:
        return {
            "test_date_raw": raw_date,
            "test_date_iso": None,
            "test_datetime_iso": None,
            "date_parse_status": "missing",
            "date_parse_rule": "missing_date",
        }

    parsed_date = None
    parsed_time = None
    status = "ok"
    rule = None

    if isinstance(date_value, datetime):
        parsed_date = date_value.date()
        parsed_time = date_value.time().replace(microsecond=0)
        rule = "excel_datetime"
    elif isinstance(date_value, date):
        parsed_date = date_value
        rule = "excel_date"
    else:
        txt = raw_date.split(" ")[0]
        for fmt, fmt_rule in (
            ("%d/%m/%Y", "DD/MM/YYYY"),
            ("%d-%m-%Y", "DD-MM-YYYY"),
            ("%Y-%m-%d", "YYYY-MM-DD"),
            ("%Y/%m/%d", "YYYY/MM/DD"),
            ("%m/%d/%Y", "MM/DD/YYYY_fallback"),
        ):
            try:
                parsed_date = datetime.strptime(txt, fmt).date()
                rule = fmt_rule
                if fmt in ("%d/%m/%Y", "%m/%d/%Y"):
                    parts = txt.replace("-", "/").split("/")
                    if len(parts) == 3 and int(parts[0]) <= 12 and int(parts[1]) <= 12:
                        status = "ambiguous"
                        rule = f"{fmt_rule}_ambiguous"
                break
            except Exception:
                pass

    if parsed_date is None:
        return {
            "test_date_raw": raw_date,
            "test_date_iso": None,
            "test_datetime_iso": None,
            "date_parse_status": "error",
            "date_parse_rule": "unparsed",
        }

    if parsed_time is None and raw_time:
        if isinstance(time_value, datetime):
            parsed_time = time_value.time().replace(microsecond=0)
        elif isinstance(time_value, time):
            parsed_time = time_value.replace(microsecond=0)
        else:
            ttxt = raw_time.split(" ")[-1]
            for tfmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
                try:
                    parsed_time = datetime.strptime(ttxt, tfmt).time()
                    break
                except Exception:
                    pass

    if parsed_time is None:
        parsed_time = time(0, 0, 0)
        dt_rule = f"{rule}; time_missing"
    else:
        dt_rule = rule

    parsed_dt = datetime.combine(parsed_date, parsed_time)
    return {
        "test_date_raw": raw_date,
        "test_date_iso": parsed_date.isoformat(),
        "test_datetime_iso": parsed_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "date_parse_status": status,
        "date_parse_rule": dt_rule,
    }


def make_hash(variant, serial, ptn, id_point, pnum, source_file):
    """Hash Ãºnico de un punto, para idempotencia."""
    key = f"{variant}|{serial}|{ptn}|{id_point}|{pnum}|{Path(source_file).name}"
    return hashlib.sha256(key.encode("utf-8")).hexdigest()


DATE_COLUMNS = {
    "test_date_raw": "VARCHAR",
    "test_date_iso": "VARCHAR",
    "test_datetime_iso": "VARCHAR",
    "date_parse_status": "VARCHAR",
    "date_parse_rule": "VARCHAR",
    "stable_point_key": "VARCHAR",
}


def ensure_schema_migrations(engine):
    """Agrega columnas nuevas en bases SQLite ya existentes."""
    with engine.begin() as con:
        cols = {row[1] for row in con.exec_driver_sql("PRAGMA table_info(test_points)")}
        for col, sql_type in DATE_COLUMNS.items():
            if col not in cols:
                con.exec_driver_sql(f"ALTER TABLE test_points ADD COLUMN {col} {sql_type}")


# ============================================================================
# ETL
# ============================================================================
def load_mapping(path):
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_effective_mapping(path, config_db=None):
    """load_mapping(path) + los parametros que el usuario agrego desde la app
    (boton Parametros > Cargar nuevos parametros), guardados en config.db sin
    tocar mapping.yaml (que es un archivo curado, con comentarios que un
    yaml.safe_dump destruiria).

    Superpone measurements[canonical][variant] = [raw_name] por cada
    custom_param registrado, sin duplicar si el mismo raw_name ya existe ahi.
    """
    import config_store as cfg

    mapping = load_mapping(path)
    db_path = config_db or cfg.CONFIG_DB
    measurements = mapping.setdefault("measurements", {})
    for canonical, raw_name, variant in cfg.list_custom_params(db_path):
        por_variante = measurements.setdefault(canonical, {})
        raw_names = por_variante.setdefault(variant, [])
        if isinstance(raw_names, str):
            raw_names = [raw_names]
            por_variante[variant] = raw_names
        if raw_name not in raw_names:
            raw_names.append(raw_name)
    return mapping


def get_identity(buf, mapping, variant):
    """Extrae los campos de identidad para esta variante."""
    out = {}
    for canon, pervar in mapping.get("identity", {}).items():
        raw = pervar.get(variant)
        if raw is None:
            out[canon] = None
            continue
        val, _unit = buf.get(norm(raw), (None, None))
        out[canon] = None if val is None else str(val).strip()
    return out


def normalize_unit(unit, mapping):
    if unit is None:
        return None
    u = str(unit).strip()
    table = mapping.get("unit_normalization", {})
    return table.get(u, u)


def add_derived_measurements(rows, variant):
    """Agrega mediciones calculadas que no existen directamente en el Buffer."""
    by_raw = {str(raw).upper(): (canon, raw, val, unit) for canon, raw, val, unit in rows}

    def add_ratio(raw_num, raw_den, out_raw="OCPR"):
        num = by_raw.get(raw_num.upper())
        den = by_raw.get(raw_den.upper())
        if not num or not den:
            return
        den_val = den[2]
        if den_val in (None, 0):
            return
        if out_raw.upper() in by_raw:
            return
        value = num[2] / den_val
        rows.append(("ocpr", out_raw, value, "ratio"))
        by_raw[out_raw.upper()] = ("ocpr", out_raw, value, "ratio")

    if variant == "1B":
        add_ratio("PS3", "PT2psia")
    elif variant == "CFM56-5A":
        add_ratio("PS3SEL", "PAMB_psi")


def get_measurements(buf, mapping, variant):
    """Devuelve lista de (canonical, raw_name, value, unit) para esta variante."""
    rows = []
    for canon, pervar in mapping.get("measurements", {}).items():
        raw_names = pervar.get(variant)
        if not raw_names:
            continue
        if isinstance(raw_names, str):
            raw_names = [raw_names]
        for raw in raw_names:
            entry = buf.get(norm(raw))
            if entry is None:
                continue  # no presente en este archivo: se omite
            val, unit = entry
            fval = to_float(val)
            if fval is None:
                continue  # valor no numerico: se omite (no se inventa)
            rows.append((canon, raw, fval, normalize_unit(unit, mapping)))
    add_derived_measurements(rows, variant)
    return rows

def process_file(session, path, mapping):
    fname = path.name
    buf = read_buffer(path)
    if buf is None:
        return ("error", "No se encontro hoja 'Buffer'", None)

    variant = detect_variant(fname, buf)
    if variant is None:
        return ("error", f"No se pudo detectar motor soportado (LEAP-1A/1B, CFM56-5A/7B): {fname}", None)

    ident = get_identity(buf, mapping, variant)
    date_info = parse_excel_date_time(ident.get("date"), ident.get("time"))
    serial = ident.get("serial_number") or "DESCONOCIDO"
    row_hash = make_hash(variant, serial, ident.get("point_test_number"),
                         ident.get("id_point"), ident.get("point_number"), fname)

    # idempotencia: Â¿ya existe este punto?
    existing = session.execute(
        select(TestPoint).where(TestPoint.row_hash == row_hash)
    ).scalar_one_or_none()
    if existing is not None:
        return ("skipped", f"Punto ya cargado (hash existe): {fname}", variant)

    # engine (crea si no existe)
    engine_type = display_engine_type(variant)
    eng = session.execute(
        select(Engine).where(Engine.serial_number == serial)
    ).scalar_one_or_none()
    if eng is None:
        eng = Engine(serial_number=serial, engine_type=engine_type)
        session.add(eng)
        session.flush()

    # test_point
    tp = TestPoint(
        engine_id=eng.id,
        variant=variant,
        serial_number=serial,
        point_test_number=ident.get("point_test_number"),
        id_point=ident.get("id_point"),
        point_type=ident.get("point_type"),
        point_number=ident.get("point_number"),
        description=ident.get("description"),
        test_date=date_info["test_date_raw"] or ident.get("date"),
        test_time=ident.get("time"),
        test_date_raw=date_info["test_date_raw"],
        test_date_iso=date_info["test_date_iso"],
        test_datetime_iso=date_info["test_datetime_iso"],
        date_parse_status=date_info["date_parse_status"],
        date_parse_rule=date_info["date_parse_rule"],
        source_file=fname,
        row_hash=row_hash,
        stable_point_key=row_hash,
        ingested_at=datetime.now(timezone.utc),
    )
    session.add(tp)
    session.flush()

    # measurements
    meas = get_measurements(buf, mapping, variant)
    for canon, raw, val, unit in meas:
        session.add(Measurement(
            point_id=tp.id, canonical=canon, raw_name=raw, value=val, unit=unit
        ))

    return ("ok", f"{len(meas)} mediciones cargadas ({engine_type})", variant)


def _list_excel_files(folder):
    return sorted(
        p for p in folder.iterdir()
        if p.suffix.lower() in (".xls", ".xlsm", ".xlsx") and not p.name.startswith("~")
    )


def run(folder, db_path, mapping_path):
    folder = Path(folder)
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)

    mapping = load_effective_mapping(mapping_path)
    engine = create_engine(f"sqlite:///{db_path}")
    Base.metadata.create_all(engine)
    ensure_schema_migrations(engine)

    files = _list_excel_files(folder)
    print(f"Archivos a procesar: {len(files)}")
    counts = {"ok": 0, "skipped": 0, "error": 0}

    with Session(engine) as session:
        for f in files:
            try:
                status, detail, _variant = process_file(session, f, mapping)
            except Exception as e:
                status, detail = "error", f"{type(e).__name__}: {e}"
            counts[status] = counts.get(status, 0) + 1
            session.add(IngestLog(
                source_file=f.name, status=status, detail=detail,
                ingested_at=datetime.now(timezone.utc),
            ))
            print(f"  [{status.upper()}] {f.name}: {detail}")
        session.commit()

    print(f"\nResumen: {counts['ok']} cargados, {counts['skipped']} omitidos, "
          f"{counts['error']} con error.")
    print(f"Base de datos: {db_path.resolve()}")


def run_sync(unloaded_dir, loaded_dir, db_path, mapping_path):
    """Carga los Exceles de unloaded_dir a db_path y mueve los que cargan
    bien (status "ok") a loaded_dir/<variante_display>/. Los duplicados
    (skipped) y los que fallan (error) se quedan en unloaded_dir.

    Devuelve un dict resumen: ok, skipped, error, moved, move_errors,
    mensajes (lista de strings por archivo, para mostrar en la UI).
    """
    import shutil

    unloaded_dir = Path(unloaded_dir)
    loaded_dir = Path(loaded_dir)
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)

    mapping = load_effective_mapping(mapping_path)
    engine = create_engine(f"sqlite:///{db_path}")
    Base.metadata.create_all(engine)
    ensure_schema_migrations(engine)

    files = _list_excel_files(unloaded_dir)
    counts = {"ok": 0, "skipped": 0, "error": 0, "moved": 0, "move_errors": 0}
    mensajes = []

    with Session(engine) as session:
        for f in files:
            try:
                status, detail, variant = process_file(session, f, mapping)
            except Exception as e:
                status, detail, variant = "error", f"{type(e).__name__}: {e}", None
            counts[status] = counts.get(status, 0) + 1
            session.add(IngestLog(
                source_file=f.name, status=status, detail=detail,
                ingested_at=datetime.now(timezone.utc),
            ))
            print(f"  [{status.upper()}] {f.name}: {detail}")

            if status == "ok" and variant is not None:
                dest_dir = loaded_dir / display_engine_type(variant)
                try:
                    dest_dir.mkdir(parents=True, exist_ok=True)
                    dest = dest_dir / f.name
                    seq = 2
                    while dest.exists():
                        dest = dest_dir / f"{f.stem} ({seq}){f.suffix}"
                        seq += 1
                    shutil.move(str(f), str(dest))
                    counts["moved"] += 1
                    mensajes.append(f"[OK] {f.name}: cargado y movido a {dest_dir.name}/")
                except Exception as e:
                    counts["move_errors"] += 1
                    mensajes.append(
                        f"[OK, NO MOVIDO] {f.name}: cargado en la base pero no se pudo "
                        f"mover ({type(e).__name__}: {e}) - muevelo manualmente a "
                        f"Loaded/{display_engine_type(variant)}/"
                    )
            else:
                mensajes.append(f"[{status.upper()}] {f.name}: {detail}")

        session.commit()

    print(f"\nResumen sync: {counts['ok']} cargados ({counts['moved']} movidos a Loaded), "
          f"{counts['skipped']} omitidos, {counts['error']} con error.")
    return {"total": len(files), "mensajes": mensajes, **counts}


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="ETL piloto LEAP -> SQLite")
    ap.add_argument("folder", help="Carpeta con archivos .xlsm de una variante")
    ap.add_argument("--db", default="data/motores.db", help="Ruta del archivo SQLite")
    ap.add_argument("--mapping", default="mapping.yaml", help="Ruta del mapping.yaml")
    args = ap.parse_args()
    run(args.folder, args.db, args.mapping)
